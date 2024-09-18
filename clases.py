import texts
import config
import sqlite3
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.dispatcher.filters.state import State, StatesGroup
import re
import openpyxl
from openpyxl.styles import Font
import os
import datetime

class DataBase:
    def __init__(self, db_name):
        """Управление базой данных"""
        self.db_name = db_name
    
    # Создания
    def create_db(self):
        """Создание БД и таблиц"""
        create_table_queries = [
            """
            CREATE TABLE IF NOT EXISTS bot_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                chat_id INTEGER UNIQUE
            );
            """,
            """
            CREATE TABLE IF NOT EXISTS users_data (
                id INTEGER PRIMARY KEY,
                user_name TEXT,
                first_name TEXT,
                last_name TEXT,
                name TEXT,
                company TEXT,
                job_title TEXT,
                email TEXT,
                phone_quest TEXT,
                agreement INTEGER DEFAULT 0,
                phone_telegram TEXT,
                FOREIGN KEY (id) REFERENCES bot_users (id)
            );
            """,
            """
            CREATE TABLE IF NOT EXISTS admin (
                id INTEGER PRIMARY KEY,
                FOREIGN KEY (id) REFERENCES bot_users (id)
            );
            """,
            """
            CREATE TABLE IF NOT EXISTS question (
                id INTEGER PRIMARY KEY,
                FOREIGN KEY (id) REFERENCES bot_users (id)
            );
            """,
             """
            CREATE TABLE IF NOT EXISTS game_register (
                id INTEGER PRIMARY KEY,
                loto_number INTEGER DEFAULT 1,
                FOREIGN KEY (id) REFERENCES bot_users (id)
            );
            """,
            """
            CREATE TABLE IF NOT EXISTS coun (
                start INTEGER DEFAULT 0,
                registration INTEGER DEFAULT 0,
                catalog INTEGER DEFAULT 0,
                price INTEGER DEFAULT 0,
                game_now INTEGER DEFAULT 0,
                game_all INTEGER DEFAULT 0,
                email_all INTEGER DEFAULT 0,
                phone_all INTEGER DEFAULT 0
            );
            """,
            """
            CREATE TABLE IF NOT EXISTS status (
                event TEXT,
                status BOOL DEFAULT 0
            );
            """
        ]

        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            for query in create_table_queries:
                cursor.execute(query)
            conn.commit()

    def create_game(self):
        """Создание строки в таблице status со значением game и 0"""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            try:
                # Проверяем, существует ли уже запись с event = 'game'
                cursor.execute("SELECT COUNT(*) FROM status WHERE event = 'game'")
                result = cursor.fetchone()
                if result[0] > 0:
                    print("Запись с event = 'game' уже существует.")
                    return  # Выходим из метода, если запись уже существует

                # Если записи нет, создаем новую запись
                query = """
                INSERT INTO status (event, status) VALUES ('game', 0);
                """
                cursor.execute(query)
                conn.commit()
                print("Запись с event = 'game' успешно создана.")
            except Exception as e:
                print(f"Ошибка при создании записи: {e}")
        

    def add_god(self, chat_id):
        """Добавление первой записи chat_id в таблицы bot_users и admin"""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            try:
                # Добавляем chat_id в таблицу bot_users
                cursor.execute("INSERT INTO bot_users (chat_id) VALUES (?)", (chat_id,))
                user_id = cursor.lastrowid  # Получаем id только что добавленной записи

                # Добавляем запись в таблицу admin с соответствующим user_id
                cursor.execute("INSERT INTO admin (id) VALUES (?)", (user_id,))
                conn.commit()
                return True  # Возвращаем True, если добавление прошло успешно
            except sqlite3.IntegrityError:
                # Если chat_id уже существует в таблице bot_users, откатываем транзакцию
                conn.rollback()
                return False  # Возвращаем False, если добавление не удалось

    def ensure_coun_table_has_record(self):
        """Проверяет, есть ли запись в таблице coun, и добавляет ее, если ее нет."""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            try:
                # Проверяем, есть ли записи в таблице coun
                cursor.execute("SELECT COUNT(*) FROM coun")
                count = cursor.fetchone()[0]
                if count == 0:
                    # Если записей нет, добавляем начальную запись
                    cursor.execute("""
                        INSERT INTO coun (start, registration, catalog, price, game_now, game_all, email_all, phone_all)
                        VALUES (0, 0, 0, 0, 0, 0, 0, 0)
                    """)
                    conn.commit()
                    print("Добавлена начальная запись в таблицу coun")
                else:
                    print("Таблица coun уже содержит записи")
            except sqlite3.Error as e:
                print(f"Ошибка при проверке/добавлении записи в таблицу coun: {e}")
                conn.rollback()
    
    # Проверки
    def chek_users(self, chat_id):
        """Проверяет есть ли обратившийся в списке пользователей"""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM bot_users WHERE chat_id = ?", (chat_id,))
            result = cursor.fetchone()
            # Возвращаем True, если пользователь найден, иначе False
            return result is not None  

    def chek_admin(self,chat_id):
        """Проверяет пользователя на админа"""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            try:
                # Находим id пользователя в таблице bot_users по chat_id
                cursor.execute("SELECT id FROM bot_users WHERE chat_id = ?", (chat_id,))
                result = cursor.fetchone()
                if result is None:
                    return False  # Пользователь не найден
                user_id = result[0]
                # Проверяем, есть ли этот id в таблице admin
                cursor.execute("SELECT COUNT(*) FROM admin WHERE id = ?", (user_id,))
                result = cursor.fetchone()
                return result[0] > 0  # Возвращаем True, если запись найдена, иначе False
            except Exception as e:
                print(f"Ошибка при проверке пользователя на админа: {e}")
                return False

    def check_guestion(self, chat_id):
        """Проверка наличия chat_id в таблицах bot_users и question"""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            # Проверяем наличие chat_id в таблице bot_users
            cursor.execute("SELECT id FROM bot_users WHERE chat_id = ?", (chat_id,))
            result = cursor.fetchone()
            if result is None:
                print(f"Пользователь с chat_id {chat_id} не найден в таблице bot_users")
                pass # Вернуть что-то другое, чтобы попросить пользователя нажать на /start
            else:
                user_id = result[0]
                # Проверяем наличие user_id в таблице question
                cursor.execute("SELECT id FROM question WHERE id = ?", (user_id,))
                result = cursor.fetchone()
                return result is not None  # Возвращаем True, если запись найдена, иначе False

    def check_phone(self, number) -> bool:
        """Проверяет корректность введенного номера телефона"""
        pattern = r'^(\+7|8)\d{10}$'
        return re.match(pattern, number) is not None

    def check_email(self, email) -> bool:
        """Проверяет корректность введенного email"""
        pattern = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
        return re.match(pattern, email) is not None

    def check_in_none(self, data: str) -> bool:
        """Входящая строка должна содержать хотя бы 2 буквы. Латинские или русские"""
        pattern = r'[a-zA-Zа-яА-Я]'
        letters = re.findall(pattern, data)
        return len(letters) >= 2

    def check_registration_game(self, param):
        """Проверяет статус доступности регистрации"""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            try:
                # Находим запись в таблице status, где event равен param
                cursor.execute("SELECT status FROM status WHERE event = ?", (param,))
                result = cursor.fetchone()
                if result is None:
                    return False  # Запись не найдена
                status = result[0]

                # Возвращаем True, если status равен 1, иначе False
                return status == 1
            except Exception as e:
                print(f"Ошибка при проверке статуса регистрации: {e}")
                return False

    def chek_game_register_id(self, chat_id):
        """Проверяет есть ли запись с id в таблице game_register"""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            try:
                # Находим запись в таблице bot_users по chat_id
                cursor.execute("SELECT id FROM bot_users WHERE chat_id = ?", (chat_id,))
                result = cursor.fetchone()
                if result is None:
                    return False  # Пользователь не найден
                user_id = result[0]

                # Проверяем, есть ли этот id в таблице game_register
                cursor.execute("SELECT COUNT(*) FROM game_register WHERE id = ?", (user_id,))
                result = cursor.fetchone()
                return result[0] > 0  # Возвращаем True, если запись найдена, иначе False
            except Exception as e:
                print(f"Ошибка при проверке записи в game_register: {e}")
                return False

    #Обновление данных в БД
    def post_new_users(self, chat_id):
        """Добавляет пользователя в БД, таблица bot_users"""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            cursor.execute("INSERT INTO bot_users (chat_id) VALUES (?)", (chat_id,))
            conn.commit()
            return True  # Возвращаем True, если вставка прошла успешно
    
    def post_users_info(self, chat_id, user_name, first_name, last_name):
        """Добавляет данные о пользователе полученные из месседж бокса в таблицу users_data"""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            try:
                # Сначала находим id пользователя в таблице bot_users
                cursor.execute("SELECT id FROM bot_users WHERE chat_id = ?", (chat_id,))
                result = cursor.fetchone()
                if result is None:
                    return False  # Пользователь не найден
                user_id = result[0]

                # Добавляем информацию о пользователе в таблицу users_data
                cursor.execute("""
                    INSERT INTO users_data (id, user_name, first_name, last_name)
                    VALUES (?, ?, ?, ?)
                """, (user_id, user_name, first_name, last_name))
                conn.commit()
                return True  # Возвращаем True, если добавление прошло успешно
            except sqlite3.Error as e:
                print(f"Ошибка при добавлении данных пользователя: {e}")
                conn.rollback()
                return False  # Возвращаем False, если добавление не удалось

    def post_users_question_data(self, chat_id, name, company, job_title, email, phone_quest):
        """Добавляет анкетные данные в таблицу users_data"""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            try:
                # Находим индекс записи в таблице bot_users по chat_id
                cursor.execute("SELECT id FROM bot_users WHERE chat_id = ?", (chat_id,))
                result = cursor.fetchone()
                if result is None:
                    return False  # Пользователь не найден
                user_id = result[0]

                # Вставляем или обновляем данные в таблице users_data
                cursor.execute("""
                    INSERT INTO users_data (id, name, company, job_title, email, phone_quest)
                    VALUES (?, ?, ?, ?, ?, ?)
                    ON CONFLICT(id) DO UPDATE SET
                    name=excluded.name,
                    company=excluded.company,
                    job_title=excluded.job_title,
                    email=excluded.email,
                    phone_quest=excluded.phone_quest
                """, (user_id, name, company, job_title, email, phone_quest))
                conn.commit()
                return True  # Возвращаем True, если вставка прошла успешно
            
            except sqlite3.Error as e:
                print(f"Ошибка при добавлении анкетных данных: {e}")
                conn.rollback()
                return False  # Возвращаем False, если вставка не удалась

    def post_telegram_phone(self, chat_id, phone):
        """Обновляет номер телефона в таблице users_data"""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            try:
                # Находим индекс записи в таблице bot_users по id
                cursor.execute("SELECT id FROM bot_users WHERE chat_id = ?", (chat_id,))
                result = cursor.fetchone()
                if result is None:
                    return False  # Пользователь не найден
                user_id = result[0]

                # Обновляем номер телефона в таблице users_data
                cursor.execute("""
                    UPDATE users_data
                    SET phone_telegram = ?
                    WHERE id = ?
                """, (phone, user_id))
                conn.commit()
                return True  # Возвращаем True, если обновление прошло успешно
            except Exception as e:
                print(f"Ошибка при обновлении номера телефона: {e}")
                return False

    def post_user_agreements(self,chat_id):
        """Меняет данные в таблице users_data, в столбце agreement. Вызывается в случае answer=1"""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            try:
                # Находим индекс записи в таблице bot_users по chat_id
                cursor.execute("SELECT id FROM bot_users WHERE chat_id = ?", (chat_id,))
                result = cursor.fetchone()
                if result is None:
                    return False  # Пользователь не найден
                user_id = result[0]

                # Обновляем значение в столбце agreement
                cursor.execute("""
                    UPDATE users_data
                    SET agreement = 1
                    WHERE id = ?
                """, (user_id,))
                conn.commit()
                return True  # Возвращаем True, если обновление прошло успешно

            except sqlite3.Error as e:
                print(f"Ошибка при обновлении данных пользователя: {e}")
                conn.rollback()
                return False  # Возвращаем False, если обновление не удалось

    def post_user_question(self, chat_id):
        """Находит id пользователя в bot_users и добавляет id пользователя в таблицу question"""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            try:
                # Находим id пользователя в таблице bot_users по chat_id
                cursor.execute("SELECT id FROM bot_users WHERE chat_id = ?", (chat_id,))
                result = cursor.fetchone()
                if result is None:
                    return False  # Пользователь не найден
                user_id = result[0]

                # Проверяем, есть ли id пользователя в таблице question
                cursor.execute("SELECT id FROM question WHERE id = ?", (user_id,))
                result = cursor.fetchone()
                if result is not None:
                    return True  # id пользователя уже есть в таблице question

                # Добавляем id пользователя в таблицу question
                cursor.execute("INSERT INTO question (id) VALUES (?)", (user_id,))
                conn.commit()
                return True  # Возвращаем True, если вставка прошла успешно
            
            except sqlite3.Error as e:
                print(f"Ошибка при добавлении id пользователя в таблицу question: {e}")
                conn.rollback()
                return False  # Возвращаем False, если вставка не удалась

    def post_counter(self, param):
        """Обращается к таблице coun и меняет значение на +1 по param"""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            try:
                # Обновляем значение счетчика
                cursor.execute(f"""
                    UPDATE coun
                    SET {param} = {param} + 1
                """)
                conn.commit()
                return True  # Возвращаем True, если обновление прошло успешно

            except sqlite3.Error as e:
                print(f"Ошибка при обновлении счетчика: {e}")
                conn.rollback()
                return False  # Возвращаем False, если обновление не удалось

    def post_new_gamer(self, chat_id, number):
        """Записывает id и number в таблицу game_register"""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            try:
                # Находим id пользователя в таблице bot_users по chat_id
                cursor.execute("SELECT id FROM bot_users WHERE chat_id = ?", (chat_id,))
                result = cursor.fetchone()
                if result is None:
                    print(f"Пользователь с chat_id {chat_id} не найден.")
                    return False  # Пользователь не найден
                user_id = result[0]

                # Вставляем запись в таблицу game_register
                cursor.execute("""
                    INSERT INTO game_register (id, loto_number)
                    VALUES (?, ?)
                """, (user_id, number))
                conn.commit()
                return True  # Возвращаем True, если вставка прошла успешно
            except Exception as e:
                print(f"Ошибка при добавлении записи в game_register: {e}")
                return False

    def post_game_status(self, event, status):
        """Меняет status по param в таблице status"""
        try:
            # Использование контекстного менеджера для подключения к базе данных
            with sqlite3.connect(self.db_name) as conn:
                cursor = conn.cursor()
                # SQL-запрос для обновления статуса
                query = f"UPDATE status SET status = ? WHERE event = ?"
                cursor.execute(query, (status, event))
                # Сохранение изменений
                conn.commit()
        except sqlite3.Error as e:
            print(f"Ошибка при обновлении статуса игры: {e}")

    def clear_game_register(self):
        """Очищает таблицу game_register"""
        try:
            with sqlite3.connect(self.db_name) as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM game_register")
                conn.commit()
                print("Таблица game_register успешно очищена.")
        except sqlite3.Error as e:
            print(f"Ошибка при очистке таблицы game_register: {e}")
        pass

    #Обработать данные
    def format_lot_number(self, number):
        """Форматирует число, используя символы из словаря lot_number"""
        # Преобразуем число в строку
        number_str = str(number)
        # Заменяем каждую цифру на соответствующий символ из словаря
        formatted_number = ''.join(texts.lot_number[int(digit)] for digit in number_str)
        return formatted_number

    #Получить данные
    def get_game_count(self):
        """Обращается к таблице game_register и возвращает int количества записей в ней"""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            try:
                # Выполняем запрос, который считает количество строк в таблице game_register
                cursor.execute("SELECT COUNT(*) FROM game_register")
                result = cursor.fetchone()
                return result[0]  # Возвращаем количество строк
            except Exception as e:
                print(f"Ошибка при подсчете записей в таблице game_register: {e}")
                return False  # Возвращаем False в случае ошибки
    
    def get_user_lot(self, chat_id):
        """Извлекает данные из таблицы game_register и соответствующие номера телефонов из таблицы users_data"""
        try:
            with sqlite3.connect(self.db_name) as conn:
                cursor = conn.cursor()

                # Находим id пользователя в таблице bot_users по chat_id
                cursor.execute("SELECT id FROM bot_users WHERE chat_id = ?", (chat_id,))
                result = cursor.fetchone()
                if result is None:
                    return None  # Пользователь не найден
                user_id = result[0]

                # Получаем данные из таблицы game_register и соответствующие номера телефонов из таблицы users_data
                cursor.execute("""
                    SELECT gr.loto_number, ud.phone_telegram
                    FROM game_register gr
                    JOIN users_data ud ON gr.id = ud.id
                    WHERE gr.id = ?
                """, (user_id,))

                result = cursor.fetchone()
                if result is None:
                    return None  # Запись не найдена

                return result  # Возвращаем кортеж (loto_number, phone_telegram)
        except sqlite3.Error as e:
            print(f"Ошибка при извлечении данных из game_register и users_data: {e}")
            return None
    
    def get_all_users(self):
        """Получает список всех chat_id"""
        with sqlite3.connect(self.db_name) as conn:
            cursor = conn.cursor()
            try:
                # Выполняем запрос к таблице bot_users, чтобы получить все chat_id
                cursor.execute("SELECT chat_id FROM bot_users")
                results = cursor.fetchall()
                # Извлекаем chat_id из результатов запроса и формируем список
                chat_ids = [result[0] for result in results]
                return chat_ids
            except Exception as e:
                print(f"Ошибка при получении списка chat_id: {e}")

    #Получаем chat_id из таблицы game_register
    def get_game_register_data(self):
        """Извлекает данные из таблицы game_register"""
        try:
            with sqlite3.connect(self.db_name) as conn:
                cursor = conn.cursor()

                # Получаем список id из таблицы game_register
                query = "SELECT id FROM game_register"
                cursor.execute(query)
                game_register_ids = [row[0] for row in cursor.fetchall()]

                # Получаем список chat_id из таблицы bot_users по id из game_register
                chat_ids = []
                for user_id in game_register_ids:
                    query = "SELECT chat_id FROM bot_users WHERE id = ?"
                    cursor.execute(query, (user_id,))
                    result = cursor.fetchone()
                    if result:
                        chat_ids.append(result[0])

                return chat_ids
        except sqlite3.Error as e:
            print(f"Ошибка при извлечении данных из game_register и bot_users: {e}")
            return []

    def get_quest_users(self, chat_id):
        """Обращается к таблице bot_users получает id пользователя по chat_id и проверяет наличие id в таблице user_data"""
        try:
            with sqlite3.connect(self.db_name) as conn:
                cursor = conn.cursor()

                # Получаем id пользователя из таблицы bot_users по chat_id
                cursor.execute("SELECT id FROM bot_users WHERE chat_id = ?", (chat_id,))
                result = cursor.fetchone()
                if result is None:
                    print(f"Пользователь с chat_id {chat_id} не найден в таблице bot_users.")
                    return False
                user_id = result[0]

                # Проверяем наличие id в таблице user_data
                cursor.execute("SELECT id FROM users_data WHERE id = ?", (user_id,))
                result = cursor.fetchone()
                if result is None:
                    print(f"Пользователь с id {user_id} не найден в таблице users_data.")
                    return False

                print(f"Пользователь с chat_id {chat_id} и id {user_id} найден в таблице users_data.")
                return True
        except sqlite3.Error as e:
            print(f"Ошибка при проверке наличия пользователя в таблицах bot_users и users_data: {e}")
            return False

    #Создание контрольной таблицы для розыгрыша 
    def create_excel_file(self, data_dict, file_path):
        """Создает Excel-файл с данными из словаря, где ключ - это loto_number, а значение - номер телефона"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Game Register"

            # Заголовки столбцов
            headers = ["Loto Number", "Phone Telegram"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)

            # Заполнение данных
            row_num = 2
            for loto_number, phone_telegram in data_dict.items():
                ws.cell(row=row_num, column=1, value=loto_number)
                ws.cell(row=row_num, column=2, value=phone_telegram)
                row_num += 1

            wb.save(file_path)
            print(f"Excel-файл успешно создан: {file_path}")
        except Exception as e:
            print(f"Ошибка при создании Excel-файла: {e}")
       
    
    def file_delete(self, file_path):
        """Отправляет файл пользователю и удаляет его, если файл существует"""
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                print(f"Файл успешно отправлен и удален: {file_path}")
            else:
                print(f"Файл не найден: {file_path}")
        except Exception as e:
            print(f"Ошибка при отправке или удалении файла: {e}")


    def get_count(self):
        """Проходит по БД и получает всю актуальную информацию по цифрам"""
        pass

class KeyBoard:
    def __init__(self):
        pass

    def get_main_menu(self):
        """Генерирует главное меню с кнопками."""
        markup = ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(KeyboardButton(texts.products), KeyboardButton(texts.prize))
        markup.add(KeyboardButton(texts.serch_stand), KeyboardButton(texts.fabric_info))
        return markup
    
    #InLineRegistration
    def get_confirmation_keyboard(self):
        "Рисует кнопки Да и Нет"
        keyboard = InlineKeyboardMarkup()
        keyboard.row(
            InlineKeyboardButton(texts.confirm_data_yes, callback_data='confirm_yes'),
            InlineKeyboardButton(texts.confirm_data_no, callback_data='confirm_no')
        )
        return keyboard
    
    def get_politic_confirmation(self):
        """Рисует кнопки Политика конфиденциальности и ПРИНЯТЬ"""
        keyboard = InlineKeyboardMarkup()
        privacy_policy_button = InlineKeyboardButton(texts.politic_message, url=texts.politic_url)
        update_db_button = InlineKeyboardButton(texts.confirm_politic, callback_data='confirm_politic')
        keyboard.add(privacy_policy_button)
        keyboard.add(update_db_button)
        return keyboard

   #Product 
    def get_product(self):
        """Формирует список кнопок для Каталога и Прайс-листа"""
        keyboard = InlineKeyboardMarkup()
        keyboard.row(
            InlineKeyboardButton(texts.get_catalog, callback_data='get_catalog'),
            InlineKeyboardButton(texts.get_price, callback_data='get_price_list')
        )
        return keyboard

    #Prize
    def get_prize_comunication(self, step):
        """Формирует интерактивные кнопки для регистрации: one - первыйшаг, two - второй"""
        if step == 'one':
            keyboard = InlineKeyboardMarkup()
            keyboard.row(
                InlineKeyboardButton(texts.prize_cont_inline, callback_data='next_step_prize'),
            )
            return keyboard
        elif step == 'two':
            keyboard = InlineKeyboardMarkup()
            keyboard.row(
                InlineKeyboardButton(texts.prize_reg_inline, callback_data='registration_final'),
            )
            return keyboard



        pass

    def get_contact_keyboard(self):
        """Клавиатура запроса номера телефона"""
        keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
        button_phone = KeyboardButton(text=texts.prize_registration, request_contact=True)
        keyboard.add(button_phone)
        return keyboard
    
    #Admin
    def get_admin_main_menu(self):
        """Рисует инлайн меню админки"""
        keyboard = InlineKeyboardMarkup()
        keyboard.row(
            InlineKeyboardButton(texts.post_all_message, callback_data='post_all_message'),
            InlineKeyboardButton(texts.get_game_status, callback_data='get_game_status')
        )
        keyboard.row(
            InlineKeyboardButton(texts.get_bot_state, callback_data='get_bot_state'),
            InlineKeyboardButton(texts.get_xml_file, callback_data='get_xml_file')
        )
        return keyboard
    
    def get_admin_message_confirm(self):
        """Принтует меню для отпарвки сообщений всем пользователям"""
        keyboard = InlineKeyboardMarkup()
        keyboard.row(
            InlineKeyboardButton(texts.confirm_data_yes, callback_data='confirm_message_yes'),
            InlineKeyboardButton(texts.confirm_data_no, callback_data='confirm_message_no')
        )
        keyboard.row(
            InlineKeyboardButton("Отмена", callback_data='confirm_message_cancel')
        )
        return keyboard

    #Game
    def game_main_menu(self):
        """Меню управления игрой"""
        keyboard = InlineKeyboardMarkup()
        keyboard.row(
            InlineKeyboardButton(texts.post_registration_status, callback_data='post_registration_status'),
            InlineKeyboardButton(texts.post_list_gamers, callback_data='post_list_gamers')
        )
        keyboard.row(
            InlineKeyboardButton("Отмена", callback_data='cancel_game_menu')
        )
        return keyboard

    def game_registr_control(self):
        """Меню управления регистрацией"""
        # Открыть регистрацию, закртыь регистрацию
        # Отмена
        keyboard = InlineKeyboardMarkup()
        keyboard.row(
            InlineKeyboardButton(texts.post_status_game_open, callback_data='post_status_game_open'),
            InlineKeyboardButton(texts.post_status_game_close, callback_data='post_status_game_close')
        )
        keyboard.row(
            InlineKeyboardButton("Отмена", callback_data='cancel_game_menu')
        )
        return keyboard

    def gamer_list_menu(self):
        """Публикует кнопки меню для управлением списками игроков"""
        keyboard = InlineKeyboardMarkup()
        keyboard.row(
            InlineKeyboardButton(texts.post_gamer_list, callback_data='post_gamer_list'),
            InlineKeyboardButton(texts.post_gamer_del, callback_data='post_gamer_del')
        )
        keyboard.row(
            InlineKeyboardButton("Отмена", callback_data='cancel_game_menu')
        )
        return keyboard


DataB = DataBase(config.DB_NAME) 

print(DataB.get_user_lot('166476724'))

        

    
    




        
